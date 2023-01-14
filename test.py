import discord, openpyxl, os, asyncio
from discord.ext import commands
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

PREFIX = os.environ['PREFIX']
TOKEN = os.environ['TOKEN']
game = discord.Game('test')
bot = commands.Bot(command_prefix='!',
                   intents=discord.Intents.all(),
                   status=discord.Status.online,
                   activity=game)

currentPath = os.getcwd()
os.chdir('C:/Users/user/Desktop/dbot')
files = os.listdir('C:/Users/user/Desktop/dbot/')
@bot.command()
async def t(ctx):
    print(currentPath)
    print("Files in %r: %s" % ('C:/Users/user/Desktop/dbot', files))
column_fist = 1
column_sec = 2

wb = load_workbook('test.xlsx')
ws = wb.active

def check(_id):
  for row in range(1, ws.max_row + 1):
    str_number = f'{ws.cell(row,1).value}'
    #    return False
    if ws.cell(row, 1).value == str(float(_id)):
      break
    else:
      if row is ws.max_row:
        return True
      continue

def checkRow():
  for row in range(2, ws.max_row + 2):
    if ws.cell(row, 1).value is None:
      return row

def input(_id):
  _row = checkRow()

  ws.cell(row=_row, column=column_fist, value=_id)
  ws.cell(row=_row, column=column_sec, value=0)

  wb.save('C:/Users/user/Desktop/dbot/test.xlsx')

@bot.command()
async def 테스트(ctx):
  #print(ctx.author.id)
  if check(ctx.author.id) is True:
    input(ctx.author.id)
    await ctx.send('suc')
  else:
    await ctx.send('fail')
bot.run(TOKEN)
