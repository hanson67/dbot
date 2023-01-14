import discord, openpyxl, os, asyncio
from discord.ext import commands
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

PREFIX = os.environ['PREFIX']
TOKEN = os.environ['TOKEN']
game = discord.Game('앙')
bot = commands.Bot(command_prefix='!',
                   intents=discord.Intents.all(),
                   status=discord.Status.online,
                   activity=game)

c_id = 1
c_score = 2
c_name = 3
c_mcname = 4

default_score = 0
wb = load_workbook('userDBX.xlsx')
ws = wb.active


def findRow(_id):
  for row in range(1, ws.max_row + 1):
    str_number = f'{ws.cell(row,1).value}'
    #    return False
    if ws.cell(row, 1).value == str(float(_id)):
      return row
    else:
      continue


def checkRow():
  for row in range(2, ws.max_row + 2):
    if ws.cell(row, 1).value is None:
      return row
      break


def signup(_id, _name, _mcname):
  _row = checkRow()

  ws.cell(row=_row, column=c_id, value=str(float(_id)))
  ws.cell(row=_row, column=c_score, value=default_score)
  ws.cell(row=_row, column=c_name, value=_name)
  ws.cell(row=_row, column=c_mcname, value=_mcname)

  wb.save('C:\Users\user\Desktop\dbot')


def checkScore(_id):
  for row in range(2, ws.max_row + 2):
    if ws.cell(row, 1).value == str(float(_id)):
      return ws.cell(row, 2).value
    else:
      continue


def checkName(_id):
  for row in range(1, ws.max_row + 1):
    str_number = f'{ws.cell(row,1).value}'
    #    return False
    if ws.cell(row, 1).value == str(float(_id)):
      break
    else:
      if row is ws.max_row:
        return True
      continue


def delete():
  ws.delete_rows(2, ws.max_row)
  wb.save('userDBX.xlsx')


def rankScore(lst):
  for n in range(len(lst)):
    for m in range(n, len(lst)):
      if lst[n][1] < lst[m][1]:
        lst[n], lst[m] = lst[m], lst[n]
  return lst


@bot.command()
async def 등록(ctx, mcname):
  #print(ctx.author.name)
  #print(ctx.author.id)
  if checkName(ctx.author.id) is True:
    signup(ctx.author.id, ctx.author.name, mcname)
#    await ctx.member.edit(nick=f'{mcname}(미정)')
    await ctx.send('등록이 완료되었습니다.')
  else:
    await ctx.send('이미 가입하셨습니다.')


@bot.command()
async def 활약도(ctx):
  if checkName(ctx.author.id):
    embed = discord.Embed(title='등록되어 있지않습니다.', description='', color=0x4432a8)
    await ctx.send(embed=embed)
  else:
    embed = discord.Embed(
      title=f'{ctx.author.name}님의 활약도 : {checkScore(ctx.author.id)}',
      description='',
      color=0x4432a8)
    await ctx.send(embed=embed)


@bot.command()
async def 소속배정(ctx, user: discord.User):
  print(user.id)


@bot.command()
async def 활약도랭킹(ctx):
  us = []
  for row in range(2, ws.max_row + 1):
    ui = []
    for col in range(1, 4):
      ui.append(ws.cell(row, col).value)
    us.append(ui)
  embed = discord.Embed(title='이달의 활약도 랭킹',
                        description='20위 까지만 표시',
                        color=0x4432a8)
  if ws.max_row > 21:
    for n in range(20):
      embed.add_field(name=f'{n+1}. {rankScore(us)[n][2]}',
                      value=f'{rankScore(us)[n][1]}',
                      inline=False)
  else:
    for n in range(ws.max_row - 1):
      embed.add_field(name=f'{n+1}. {rankScore(us)[n][2]}',
                      value=f'{rankScore(us)[n][1]}',
                      inline=False)
  await ctx.send(embed=embed)


@bot.command()
async def 활약도추가(ctx, i, *, user: discord.User):
  if ctx.message.author.guild_permissions.administrator:
    if checkName(user.id):
      await ctx.send("등록되지 않는 사용자입니다.")
    else:
      if int(f'{i}') < -100000 or int(f'{i}') > 100000:
        embed = discord.Embed(title='숫자 범위가 너무 큽니다.',
                              description='',
                              color=0x4432a8)
        await ctx.send(embed=embed)
      else:
        i = int(f'{i}')
        print(type(ws.cell(findRow(user.id), c_score).value), type(i))
        ws.cell(findRow(user.id), c_score).value += i
        embed = discord.Embed(
          title=f'{user.name}님의 활약도 : {checkScore(user.id)}',
          description='',
          color=0x4432a8)
        await ctx.send(embed=embed)


@bot.command()
async def reset(ctx):
  if ctx.message.author.guild_permissions.administrator:
    delete()


@bot.command()
async def hello(ctx):
  await ctx.send(f'{ctx.author.mention}님 안녕하세요!')


bot.run(TOKEN)